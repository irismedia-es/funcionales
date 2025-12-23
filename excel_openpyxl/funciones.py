import openpyxl
from openpyxl.utils import get_column_letter

def crear_planilla_notas(nombre_archivo="calificaciones.xlsx"):
    """
    Crea una planilla de notas para aplicar fórmulas.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notas Finales"
    
    # Encabezados
    ws.append(["Estudiante", "Examen 1", "Examen 2", "Examen 3", "Promedio", "Estado"])
    
    datos = [
        ["Juan Pérez", 85, 90, 88],
        ["Maria García", 70, 65, 72],
        ["Carlos Lopez", 45, 50, 40],
        ["Ana Torres", 95, 98, 100],
    ]
    
    for row_data in datos:
        ws.append(row_data)
        
    return wb, ws, nombre_archivo

def agregar_formulas(ws):
    """
    Agrega fórmulas de Excel (SUM, AVERAGE, IF) a las celdas.
    Nota: OpenPyXL escribe la cadena de la fórmula. Excel la calcula al abrir el archivo.
    """
    print("--- Agregando fórmulas ---")
    
    # Asumiendo que tenemos encabezados en fila 1, datos desde fila 2 hasta max_row
    for row in range(2, ws.max_row + 1):
        # Las columnas de notas son B(2), C(3), D(4)
        # Queremos Promedio en columna E(5) y Estado en F(6)
        
        # Fórmula PROMEDIO (AVERAGE)
        # =AVERAGE(B2:D2)
        celda_promedio = ws.cell(row=row, column=5)
        rango_notas = f"B{row}:D{row}"
        celda_promedio.value = f"=AVERAGE({rango_notas})"
        
        # Fórmula SI (IF) para ver si aprobó
        # =IF(E2>=60, "Aprobado", "Reprobado")
        celda_estado = ws.cell(row=row, column=6)
        coord_promedio = f"E{row}"
        celda_estado.value = f'=IF({coord_promedio}>=60, "Aprobado", "Reprobado")'
        
    print(f"Fórmulas aplicadas a {ws.max_row - 1} estudiantes.")

def agregar_totales(ws):
    """
    Agrega una fila de totales al final.
    """
    row_total = ws.max_row + 1
    
    ws.cell(row=row_total, column=1, value="Promedio General")
    
    # Calcular promedio de cada columna de examen
    # Columnas 2, 3, 4
    for col in range(2, 5):
        letra_col = get_column_letter(col)
        rango = f"{letra_col}2:{letra_col}{row_total-1}"
        ws.cell(row=row_total, column=col).value = f"=AVERAGE({rango})"
        
    print("Fila de promedios generales agregada.")

if __name__ == "__main__":
    wb, ws, nombre = crear_planilla_notas()
    
    # 1. Agregar fórmulas por fila
    agregar_formulas(ws)
    
    # 2. Agregar fórmulas de resumen al final
    agregar_totales(ws)
    
    wb.save(nombre)
    print(f"Archivo '{nombre}' generado con fórmulas. ¡Ábrelo en Excel para ver los cálculos!")
