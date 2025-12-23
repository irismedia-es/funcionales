import openpyxl

def preparar_datos_busqueda(nombre_archivo="datos_busqueda.xlsx"):
    """
    Crea un archivo con estructura semi-estructurada para probar la búsqueda.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Info Corporativa"
    
    # Simular una estructura tipo formulario o reporte
    ws["A1"] = "REPORTE CONFIDENCIAL"
    ws["A3"] = "Nombre de la Empresa:"
    ws["B3"] = "TechSolutions Inc."
    
    ws["A5"] = "Fecha del Reporte:"
    ws["B5"] = "2024-05-20"
    
    ws["A7"] = "CEO:"
    ws["B7"] = "Ana Sánchez"
    
    # Algunos datos tabulares más abajo
    ws["A10"] = "Métricas Clave"
    ws["A11"] = "KPI"
    ws["B11"] = "Valor"
    ws["A12"] = "ROI"
    ws["B12"] = "15%"
    ws["A13"] = "NPS"
    ws["B13"] = 78
    
    wb.save(nombre_archivo)
    return nombre_archivo

def extraer_celda_fija(ws, celda):
    """
    Extrae un campo de una posición conocida (ej. A5).
    """
    valor = ws[celda].value
    print(f"Valor fijo en {celda}: {valor}")
    return valor

def buscar_campo_por_etiqueta(ws, etiqueta_buscar, offset_col=1):
    """
    Busca una celda con un texto específico (ej. 'Company') y devuelve
    el valor de la celda a su derecha (u otro offset).
    """
    print(f"\nBuscando etiqueta: '{etiqueta_buscar}'...")
    
    encontrado = False
    
    # Iterar sobre todas las celdas usadas
    for row in ws.iter_rows():
        for cell in row:
            # Comparamos ignorando mayúsculas/minúsculas para ser más robustos
            if cell.value and str(cell.value).strip().lower() == etiqueta_buscar.lower():
                # Encontramos la etiqueta, ahora obtenemos el valor relativo
                # row=cell.row, column=cell.column + offset_col
                valor_objetivo = ws.cell(row=cell.row, column=cell.column + offset_col).value
                print(f" -> ¡Encontrado en {cell.coordinate}! Valor asociado: '{valor_objetivo}'")
                return valor_objetivo
                
    if not encontrado:
        print(" -> Etiqueta no encontrada.")
        return None

if __name__ == "__main__":
    archivo = preparar_datos_busqueda()
    wb = openpyxl.load_workbook(archivo, data_only=True)
    ws = wb.active
    
    # 1. Extracción directa
    print("--- Extracción Directa ---")
    extraer_celda_fija(ws, "B3")
    
    # 2. Búsqueda por contenido
    print("--- Búsqueda Inteligente ---")
    
    # Ejemplo 1: Buscar CEO
    ceo = buscar_campo_por_etiqueta(ws, "CEO:")
    
    # Ejemplo 2: Buscar ROI
    roi = buscar_campo_por_etiqueta(ws, "ROI")
    
    # Ejemplo 3: Buscar algo que no existe
    buscar_campo_por_etiqueta(ws, "Dirección")
