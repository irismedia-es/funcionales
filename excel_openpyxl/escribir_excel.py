import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def crear_nuevo_excel(nombre_archivo):
    """
    Crea un nuevo workbook desde cero y lo guarda.
    """
    print(f"--- Creando nuevo archivo: {nombre_archivo} ---")
    
    wb = openpyxl.Workbook()
    
    # Al crear un WB nuevo, siempre hay una hoja activa por defecto
    ws = wb.active
    ws.title = "Reporte Mensual"
    
    return wb, ws

def anadir_encabezados(ws):
    """
    Añade encabezados y les aplica estilo.
    """
    headers = ["Mes", "Departamento", "Ingresos", "Gastos", "Beneficio"]
    ws.append(headers)
    
    # Dar estilo a la fila de encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

def escribir_datos(ws):
    """
    Escribe varias filas de datos.
    """
    data = [
        ["Enero", "Marketing", 5000, 2000, 3000],
        ["Enero", "Ventas", 12000, 8000, 4000],
        ["Febrero", "Marketing", 5500, 2200, 3300],
        ["Febrero", "Ventas", 11000, 7500, 3500],
    ]
    
    for row in data:
        ws.append(row)
    
    print(f"Se han añadido {len(data)} filas de datos.")

def anadir_hoja_resumen(wb):
    """
    Crea una segunda hoja adicional.
    """
    ws2 = wb.create_sheet(title="Resumen Ejecutivo")
    ws2["A1"] = "Resumen del Año"
    ws2["A1"].font = Font(size=14, bold=True)
    ws2.merge_cells("A1:C1") # Combinar celdas
    print("Hoja 'Resumen Ejecutivo' creada.")

def escribir_celda_directa(ws):
    """
    Escribe valores en celdas específicas usando coordenadas.
    """
    # Escribir en una celda específica
    ws["G2"] = "Nota:"
    ws["H2"] = "Datos preliminares"
    
    # Usando row/column índices (1-based)
    ws.cell(row=3, column=7, value="Revisado por:")
    ws.cell(row=3, column=8, value="Admin")

if __name__ == "__main__":
    archivo_salida = "reporte_financiero.xlsx"
    
    # 1. Crear WB
    workbook, hoja_principal = crear_nuevo_excel(archivo_salida)
    
    # 2. Estructura y datos
    anadir_encabezados(hoja_principal)
    escribir_datos(hoja_principal)
    escribir_celda_directa(hoja_principal)
    
    # 3. Añadir otra hoja
    anadir_hoja_resumen(workbook)
    
    # 4. Guardar
    workbook.save(archivo_salida)
    print(f"¡Guardado exitosamente en {archivo_salida}!")
