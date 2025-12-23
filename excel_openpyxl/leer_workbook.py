import openpyxl
import os

def crear_datos_prueba(nombre_archivo="datos_prueba.xlsx"):
    """
    Crea un archivo de Excel de prueba si no existe.
    Esto es solo para asegurar que el ejemplo funcione.
    """
    if os.path.exists(nombre_archivo):
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"
    
    # Encabezados
    headers = ["ID", "Producto", "Categoria", "Precio", "Cantidad"]
    ws.append(headers)
    
    # Datos
    datos = [
        [101, "Laptop", "Electronica", 1200, 5],
        [102, "Mouse", "Electronica", 25, 50],
        [103, "Silla", "Muebles", 150, 20],
        [104, "Teclado", "Electronica", 45, 30],
        [105, "Mesa", "Muebles", 300, 10],
    ]
    
    for row in datos:
        ws.append(row)
        
    wb.save(nombre_archivo)
    print(f"Archivo de prueba '{nombre_archivo}' creado automáticamente.")

def leer_workbook(ruta_archivo):
    """
    Abre y analiza la estructura básica de un workbook.
    """
    print(f"--- Abriendo: {ruta_archivo} ---")
    
    # Cargar el workbook
    # data_only=True lee los valores calculados de las fórmulas, no la fórmula en sí
    wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
    
    # Ver las hojas disponibles
    print(f"Hojas encontradas: {wb.sheetnames}")
    
    # Seleccionar una hoja específica
    if "Ventas" in wb.sheetnames:
        sheet = wb["Ventas"]
    else:
        sheet = wb.active # La hoja activa por defecto
        
    print(f"\nAnalizando hoja: '{sheet.title}'")
    print(f"Dimensiones: {sheet.dimensions}") # Ejemplo: A1:E6
    print(f"Número de filas: {sheet.max_row}")
    print(f"Número de columnas: {sheet.max_column}")
    
    return wb, sheet

def iterar_filas(sheet):
    """
    Muestra cómo iterar sobre las filas de una hoja.
    """
    print("\n--- Iterando filas ---")
    
    # values_only=True devuelve solo los valores, no los objetos Cell
    # min_row=2 salta los encabezados
    for row in sheet.iter_rows(min_row=1, max_row=4, values_only=True):
        print(row)

def leer_celda_especifica(sheet, celda):
    """
    Lee el valor de una celda específica (ej. 'B3').
    """
    valor = sheet[celda].value
    print(f"\nValor en {celda}: {valor}")

if __name__ == "__main__":
    archivo_excel = "ventas_ejemplo.xlsx"
    crear_datos_prueba(archivo_excel)
    
    # 1. Abrir y analizar
    workbook, hoja = leer_workbook(archivo_excel)
    
    # 2. Iterar datos
    iterar_filas(hoja)
    
    # 3. Leer celda directa
    leer_celda_especifica(hoja, "B2")
